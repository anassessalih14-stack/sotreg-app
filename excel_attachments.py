from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Sequence, Tuple, Optional
import tempfile
import pandas as pd
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from calc_engine import calculate_month_to_excel


THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")


def _fmt2(x) -> str:
    try:
        return f"{float(x):.2f}"
    except Exception:
        return "" if x is None else str(x)


def _norm_vehicle_label(v: str) -> str:
    s = (v or "").strip()
    if s.upper() == "MINICAR 30P" or s == "minicar 30P" or s.lower() == "minicar 30p":
        return "minicar 30p"
    return s


def _fuse_minicar_social_outputs(df: pd.DataFrame, provider_col="Prestataire", veh_col="Type véhicule") -> pd.DataFrame:
    if df is None or df.empty or provider_col not in df.columns or veh_col not in df.columns:
        return df
    df = df.copy()
    prov = df[provider_col].astype(str).str.upper().str.strip()
    veh = df[veh_col].astype(str).str.upper()
    mask = prov.eq("STCR") & veh.str.contains("MINICAR") & veh.str.contains("SOCIAL")
    if mask.any():
        df.loc[mask, veh_col] = "minicar 30p"
    df[veh_col] = df[veh_col].astype(str).apply(_norm_vehicle_label)
    return df


def _read_month_sheets(db_path: str, period: str) -> dict[str, pd.DataFrame]:
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "tmp.xlsx"
        calculate_month_to_excel(db_path, period, tmp)
        with pd.ExcelFile(tmp) as xl:
            return {name: xl.parse(name) for name in xl.sheet_names}


def _style_table(ws, start_row: int, start_col: int, nrows: int, ncols: int, header_rows: int = 1):
    for r in range(start_row, start_row + nrows):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if r < start_row + header_rows:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
    # set reasonable column widths
    for c in range(start_col, start_col + ncols):
        ws.column_dimensions[get_column_letter(c)].width = 22 if c == start_col else 18


def _write_title(ws, entity: str, period: str, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Attachement de Transport du Personnel").font = Font(bold=True, size=16)
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, f"Entité: {entity}    Période: {period}").font = Font(bold=True, size=11)
    ws.cell(row, 1).alignment = Alignment(horizontal="left")
    return row + 2


def _provider_block_rows(entity: str, period: str, provider: str, df_ec: pd.DataFrame) -> Tuple[List[str], List[List[str]], List[List[str]]]:
    # Columns similar to your excel: Période, Véhicule, Itinéraires, Montant(HT)
    header = ["Période", "Véhicule", "Itinéraires", "Montant (HT)"]
    rows = []
    per_label = period
    for _, r in df_ec.iterrows():
        veh = str(r.get("Type véhicule","")).strip()
        age = str(r.get("Age","")).strip()
        if age and age.lower() != "nan":
            veh = f"{veh} {age}"
        veh = _norm_vehicle_label(veh)
        itin = str(r.get("Circuit","")).strip()
        amt = float(r.get("Facturation", 0) or 0)
        if amt == 0:
            continue
        rows.append([per_label, veh, itin, _fmt2(amt)])

    total_ht = sum(float(x[3]) for x in rows) if rows else 0.0
    fees = round(total_ht * 0.10, 2)
    tva = round((total_ht + fees) * 0.10, 2)
    ttc = round(total_ht + fees + tva, 2)
    totals = [
        ["", "", "TOTAL HORS TAXE", _fmt2(total_ht)],
        ["", "", "10% (Frais de Gestion)", _fmt2(fees)],
        ["", "", "TVA 10%", _fmt2(tva)],
        ["", "", "TOTAL A PAYER T.T.C", _fmt2(ttc)],
    ]
    return header, rows, totals


def _sotreg_block_rows(entity: str, period: str, df_sot: pd.DataFrame) -> Tuple[List[str], List[List[str]], List[List[str]]]:
    # Based on sotreg_lines: only autocar/minibus; table: Itinéraires, Total KM, Taux, Montant(HT)
    header = ["ITINERAIRES", "TOTAL KM PARCOURU", "TAUX", "Montant (HT)"]
    rows = []

    if df_sot.empty:
        total_ht = 0.0
        totals = [["", "", "TOTAL HORS TAXE", _fmt2(0)], ["", "", "TVA 20%", _fmt2(0)], ["", "", "TOTAL A PAYER", _fmt2(0)]]
        return header, rows, totals

    veh_col = next((c for c in ["Type véhicule","Véhicule","vehicle_type"] if c in df_sot.columns), None)
    circ_col = next((c for c in ["Circuit","Itinéraires","Itineraires"] if c in df_sot.columns), None)
    km_col = next((c for c in ["KM total","KM facturé","KM","Km Réalisée","Km Réalisé","Km Réalisé "] if c in df_sot.columns), None)
    amt_col = next((c for c in ["Frais KM","Frais de km","Montant Kilométrage","Montant (HT)","Facture","Facturation","Facture total"] if c in df_sot.columns), None)

    if not veh_col or not circ_col or not amt_col:
        total_ht = 0.0
        totals = [["", "", "TOTAL HORS TAXE", _fmt2(0)], ["", "", "TVA 20%", _fmt2(0)], ["", "", "TOTAL A PAYER", _fmt2(0)]]
        return header, rows, totals

    df = df_sot.copy()
    if "Entité" in df.columns:
        df = df[df["Entité"].astype(str)==entity].copy()
    # billing numeric
    df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0.0)
    df = df[df[amt_col] != 0].copy()
    # only autocar/minibus
    vlow = df[veh_col].astype(str).str.lower().str.strip()
    df = df[vlow.isin(["autocar","minibus"])].copy()

    if km_col and km_col in df.columns:
        df[km_col] = pd.to_numeric(df[km_col], errors="coerce").fillna(0.0)
        grp = df.groupby(circ_col, dropna=False).agg({km_col:"sum", amt_col:"sum"}).reset_index()
        for _, r in grp.iterrows():
            km = float(r[km_col])
            amt = float(r[amt_col])
            taux = (amt/km) if km else 0.0
            rows.append([str(r[circ_col]), _fmt2(km), _fmt2(taux) if km else "", _fmt2(amt)])
        total_ht = float(grp[amt_col].sum()) if not grp.empty else 0.0
    else:
        grp = df.groupby(circ_col, dropna=False).agg({amt_col:"sum"}).reset_index()
        for _, r in grp.iterrows():
            rows.append([str(r[circ_col]), "", "", _fmt2(float(r[amt_col]))])
        total_ht = float(grp[amt_col].sum()) if not grp.empty else 0.0

    tva = round(total_ht * 0.20, 2)
    total_pay = round(total_ht + tva, 2)
    totals = [
        ["", "", "TOTAL HORS TAXE", _fmt2(total_ht)],
        ["", "", "TVA 20%", _fmt2(tva)],
        ["", "", "TOTAL A PAYER", _fmt2(total_pay)],
    ]
    return header, rows, totals


def generate_provider_attachment_excel(db_path: str | Path, period: str, entity: str, provider: str, out_xlsx: str | Path) -> Path:
    db_path = str(db_path)
    out_xlsx = Path(out_xlsx)
    sheets = _read_month_sheets(db_path, period)

    wb = Workbook()
    ws = wb.active
    ws.title = "ATTACHEMENT"

    r = _write_title(ws, entity, period, 1)

    if provider.strip().upper() == "SOTREG":
        df_sot = sheets.get("sotreg_lines", pd.DataFrame())
        header, rows, totals = _sotreg_block_rows(entity, period, df_sot)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, "ANNEXE - Facture (SOTREG)").font = Font(bold=True, size=13)
        r += 2
        # write table
        for j, h in enumerate(header, 1):
            ws.cell(r, j, h)
        for i, row in enumerate(rows + totals, 1):
            for j, val in enumerate(row, 1):
                ws.cell(r+i, j, val)
        _style_table(ws, r, 1, 1 + len(rows) + len(totals), len(header), header_rows=1)
    else:
        ec = sheets.get("entity_circuit", pd.DataFrame())
        ec = _fuse_minicar_social_outputs(ec)
        if not ec.empty:
            ec = ec[(ec["Entité"].astype(str)==entity) & (ec["Prestataire"].astype(str)==provider)].copy()
            ec["Facturation"] = pd.to_numeric(ec.get("Facturation",0), errors="coerce").fillna(0.0)
            ec = ec[ec["Facturation"]!=0].copy()
        header, rows, totals = _provider_block_rows(entity, period, provider, ec if not ec.empty else pd.DataFrame())
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, f"ANNEXE - Facture ({provider})").font = Font(bold=True, size=13)
        r += 2
        for j, h in enumerate(header, 1):
            ws.cell(r, j, h)
        for i, row in enumerate(rows + totals, 1):
            for j, val in enumerate(row, 1):
                ws.cell(r+i, j, val)
        _style_table(ws, r, 1, 1 + len(rows) + len(totals), len(header), header_rows=1)

    wb.save(out_xlsx)
    return out_xlsx


def generate_global_attachment_excel(db_path: str | Path, period: str, entity: str, out_xlsx: str | Path) -> Path:
    db_path = str(db_path)
    out_xlsx = Path(out_xlsx)
    sheets = _read_month_sheets(db_path, period)

    wb = Workbook()
    ws = wb.active
    ws.title = "ATTACHEMENT_GLOBAL"

    r = _write_title(ws, entity, period, 1)

    # SOTREG block first if exists
    df_sot = sheets.get("sotreg_lines", pd.DataFrame())
    if not df_sot.empty and "Entité" in df_sot.columns and (df_sot["Entité"].astype(str)==entity).any():
        header, rows, totals = _sotreg_block_rows(entity, period, df_sot)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, "ANNEXE - Facture (SOTREG)").font = Font(bold=True, size=13)
        r += 2
        for j, h in enumerate(header, 1):
            ws.cell(r, j, h)
        for i, row in enumerate(rows + totals, 1):
            for j, val in enumerate(row, 1):
                ws.cell(r+i, j, val)
        _style_table(ws, r, 1, 1 + len(rows) + len(totals), len(header), header_rows=1)
        r = r + 2 + len(rows) + len(totals) + 2

    # Other providers from entity_circuit
    ec = sheets.get("entity_circuit", pd.DataFrame())
    ec = _fuse_minicar_social_outputs(ec)
    if not ec.empty:
        ec_ent = ec[ec["Entité"].astype(str)==entity].copy()
        providers = sorted({str(p) for p in ec_ent["Prestataire"].dropna().unique()})
        for p in providers:
            if p.strip().upper() == "SOTREG":
                continue
            dfp = ec_ent[ec_ent["Prestataire"].astype(str)==p].copy()
            dfp["Facturation"] = pd.to_numeric(dfp.get("Facturation",0), errors="coerce").fillna(0.0)
            dfp = dfp[dfp["Facturation"]!=0].copy()
            header, rows, totals = _provider_block_rows(entity, period, p, dfp)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.cell(r, 1, f"ANNEXE - Facture ({p})").font = Font(bold=True, size=13)
            r += 2
            for j, h in enumerate(header, 1):
                ws.cell(r, j, h)
            for i, row in enumerate(rows + totals, 1):
                for j, val in enumerate(row, 1):
                    ws.cell(r+i, j, val)
            _style_table(ws, r, 1, 1 + len(rows) + len(totals), len(header), header_rows=1)
            r = r + 2 + len(rows) + len(totals) + 2

    wb.save(out_xlsx)
    return out_xlsx


def generate_month_excel_with_fusion(db_path: str | Path, period: str, out_xlsx: str | Path) -> Path:
    """Generate the month excel from calc_engine, then fuse STCR minicar social -> minicar 30p in entity_circuit sheet."""
    db_path = str(db_path)
    out_xlsx = Path(out_xlsx)

    # generate raw
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "raw.xlsx"
        calculate_month_to_excel(db_path, period, tmp)
        sheets = {}
        with pd.ExcelFile(tmp) as xl:
            for name in xl.sheet_names:
                sheets[name] = xl.parse(name)

    if "entity_circuit" in sheets:
        ec = _fuse_minicar_social_outputs(sheets["entity_circuit"])
        ec["Type véhicule"] = ec["Type véhicule"].astype(str).str.replace("minicar 30P","minicar 30p", regex=False)
        group_cols = [c for c in ["Entité","Circuit","Prestataire","Type véhicule","Age"] if c in ec.columns]
        agg_map = {c:"sum" for c in ec.columns if c not in group_cols}
        if group_cols:
            ec = ec.groupby(group_cols, dropna=False, as_index=False).agg(agg_map)
        sheets["entity_circuit"] = ec

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)

    return out_xlsx


# -----------------------
# Overrides for v31
# - Option A: 2 sheets in attachment (ANNEXE + DETAIL_VERIFICATION)
# - Merge duplicate rows (same period/veh/itin) e.g. minicar 30p repeated
# - Remove 'nan' strings in labels
# -----------------------

def _provider_block_rows(entity: str, period: str, provider: str, df_ec: pd.DataFrame) -> Tuple[List[str], List[List[str]], List[List[str]]]:
    header = ["Période", "Véhicule", "Itinéraires", "Montant (HT)"]
    rows_raw = []
    per_label = period

    if df_ec is None or df_ec.empty:
        total_ht = 0.0
        fees = round(total_ht * 0.10, 2)
        tva = round((total_ht + fees) * 0.10, 2)
        ttc = round(total_ht + fees + tva, 2)
        totals = [
            ["", "", "TOTAL HORS TAXE", _fmt2(total_ht)],
            ["", "", "10% (Frais de Gestion)", _fmt2(fees)],
            ["", "", "TVA 10%", _fmt2(tva)],
            ["", "", "TOTAL A PAYER T.T.C", _fmt2(ttc)],
        ]
        return header, [], totals

    for _, r in df_ec.iterrows():
        veh = str(r.get("Type véhicule","")).strip()
        age = str(r.get("Age","")).strip()
        if age and age.lower() != "nan":
            veh = f"{veh} {age}"
        veh = _norm_vehicle_label(veh)
        veh = veh.replace(" nan","").replace(" NAN","").strip()
        itin = str(r.get("Circuit","")).strip()
        itin = itin.replace("nan","").strip()
        amt = float(r.get("Facturation", 0) or 0)
        if amt == 0:
            continue
        rows_raw.append([per_label, veh, itin, amt])

    if rows_raw:
        df = pd.DataFrame(rows_raw, columns=["Période","Véhicule","Itinéraires","Montant"])
        df["Véhicule"] = df["Véhicule"].astype(str).str.strip()
        df["Itinéraires"] = df["Itinéraires"].astype(str).str.strip()
        df = df.groupby(["Période","Véhicule","Itinéraires"], as_index=False)["Montant"].sum()
        rows = [[r["Période"], r["Véhicule"], r["Itinéraires"], _fmt2(r["Montant"])] for _, r in df.iterrows()]
        total_ht = float(df["Montant"].sum())
    else:
        rows = []
        total_ht = 0.0

    fees = round(total_ht * 0.10, 2)
    tva = round((total_ht + fees) * 0.10, 2)
    ttc = round(total_ht + fees + tva, 2)
    totals = [
        ["", "", "TOTAL HORS TAXE", _fmt2(total_ht)],
        ["", "", "10% (Frais de Gestion)", _fmt2(fees)],
        ["", "", "TVA 10%", _fmt2(tva)],
        ["", "", "TOTAL A PAYER T.T.C", _fmt2(ttc)],
    ]
    return header, rows, totals


def _build_detail_sheet_for_provider(sheets: dict, period: str, entity: str, provider: str) -> pd.DataFrame:
    if provider.strip().upper() == "SOTREG":
        df = sheets.get("sotreg_lines", pd.DataFrame()).copy()
        if df.empty:
            return df
        if "Entité" in df.columns:
            df = df[df["Entité"].astype(str) == entity].copy()
        # remove zeros
        amt_col = next((c for c in ["Facture", "Facturation", "Frais KM", "Montant Kilométrage", "TOTAL HT", "Total Facture", "TOTAL HT "] if c in df.columns), None)
        if amt_col:
            df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0.0)
            df = df[df[amt_col] != 0].copy()
        return df

    df = sheets.get("detail_prestataire", pd.DataFrame()).copy()
    if df.empty:
        return df

    if "Period" in df.columns:
        df = df[df["Period"].astype(str) == period].copy()
    if "Entité" in df.columns:
        df = df[df["Entité"].astype(str) == entity].copy()
    if "Prestataire" in df.columns:
        df = df[df["Prestataire"].astype(str) == provider].copy()

    # Fuse minicar social -> minicar 30p for STCR also in details
    if "Prestataire" in df.columns and "Type véhicule" in df.columns:
        df = _fuse_minicar_social_outputs(df, provider_col="Prestataire", veh_col="Type véhicule")

    amt_col = next((c for c in ["Facture global", "Facture", "Facturation", "Facture total", "Total Facture"] if c in df.columns), None)
    if amt_col:
        df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0.0)
        df = df[df[amt_col] != 0].copy()

    # remove literal nan strings
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].astype(str).replace("nan","").replace("None","").str.strip()
    return df


def _write_df_to_sheet(ws, df: pd.DataFrame, title: str, start_row: int = 1):
    ws.cell(start_row, 1, title).font = Font(bold=True, size=14)
    r = start_row + 2
    if df is None or df.empty:
        ws.cell(r, 1, "Aucune donnée").font = Font(italic=True)
        return

    # headers
    for j, col in enumerate(df.columns, 1):
        ws.cell(r, j, str(col))
    # rows
    for i, (_, row) in enumerate(df.iterrows(), 1):
        for j, col in enumerate(df.columns, 1):
            val = row[col]
            if isinstance(val, (int, float)) and not (val != val):  # not NaN
                ws.cell(r+i, j, float(val))
            else:
                s = "" if val is None else str(val)
                ws.cell(r+i, j, "" if s.lower() == "nan" else s)

    _style_table(ws, r, 1, 1 + len(df), len(df.columns), header_rows=1)


def generate_provider_attachment_excel(db_path: str | Path, period: str, entity: str, provider: str, out_xlsx: str | Path) -> Path:
    db_path = str(db_path)
    out_xlsx = Path(out_xlsx)
    sheets = _read_month_sheets(db_path, period)

    wb = Workbook()
    ws = wb.active
    ws.title = "ANNEXE"

    r = _write_title(ws, entity, period, 1)

    if provider.strip().upper() == "SOTREG":
        df_sot = sheets.get("sotreg_lines", pd.DataFrame())
        header, rows, totals = _sotreg_block_rows(entity, period, df_sot)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, "ANNEXE - Facture (SOTREG)").font = Font(bold=True, size=13)
        r += 2
        for j, h in enumerate(header, 1):
            ws.cell(r, j, h)
        for i, row in enumerate(rows + totals, 1):
            for j, val in enumerate(row, 1):
                ws.cell(r+i, j, val)
        _style_table(ws, r, 1, 1 + len(rows) + len(totals), len(header), header_rows=1)
    else:
        ec = sheets.get("entity_circuit", pd.DataFrame())
        ec = _fuse_minicar_social_outputs(ec)
        if not ec.empty:
            ec = ec[(ec["Entité"].astype(str)==entity) & (ec["Prestataire"].astype(str)==provider)].copy()
            ec["Facturation"] = pd.to_numeric(ec.get("Facturation",0), errors="coerce").fillna(0.0)
            ec = ec[ec["Facturation"]!=0].copy()
        header, rows, totals = _provider_block_rows(entity, period, provider, ec if not ec.empty else pd.DataFrame())
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, f"ANNEXE - Facture ({provider})").font = Font(bold=True, size=13)
        r += 2
        for j, h in enumerate(header, 1):
            ws.cell(r, j, h)
        for i, row in enumerate(rows + totals, 1):
            for j, val in enumerate(row, 1):
                ws.cell(r+i, j, val)
        _style_table(ws, r, 1, 1 + len(rows) + len(totals), len(header), header_rows=1)

    # Option A: DETAILS in second sheet
    ws2 = wb.create_sheet("DETAIL_VERIFICATION")
    df_detail = _build_detail_sheet_for_provider(sheets, period, entity, provider)
    _write_df_to_sheet(ws2, df_detail, "Détail de vérification", start_row=1)

    wb.save(out_xlsx)
    return out_xlsx
